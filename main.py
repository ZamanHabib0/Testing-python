from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
from fastapi.responses import StreamingResponse
from io import BytesIO
from fastapi.middleware.cors import CORSMiddleware
from openpyxl.styles import Border, Side
import urllib
import os
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import Font

app = FastAPI()


@app.get("/")
async def health_check():
    return {"message": "FastAPI is working!"}


app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class TaqseemkabadDetail(BaseModel):
    malikName: str
    raqbha: str
    kila: str
    mustatil: str


class TaqseemkabadEntry(BaseModel):
    details: list[TaqseemkabadDetail]


class TaqseemSaPhalaItem(BaseModel):
    malikName: str
    totalRaqbha: str
    kilaNumber: str
    mustatil: str


class RequestData(BaseModel):
    tehsil : str
    chak : str
    district : str
    khata : str
    # mizan: str
    taqseemkabad: list[TaqseemkabadEntry]
    taqseemSaPhala: list[TaqseemSaPhalaItem]

def convert_to_sqft(area):
    """ کنال-مرلہ-فٹ کو مربع فٹ میں تبدیل کریں """
    kanal, marla, feet = map(int, area.split('-'))
    total_sqft = (kanal * 20 * 272) + (marla * 272) + feet
    return int(total_sqft)  # 👈 Return as integer


def convert_from_sqft(total_sqft):
    """ مربع فٹ کو واپس کنال-مرلہ-فٹ میں تبدیل کریں """
    kanal = int(total_sqft // (20 * 272))
    remaining_sqft = total_sqft % (20 * 272)
    marla = int(remaining_sqft // 272)
    feet = round(remaining_sqft % 272)  # فٹ کو صحیح سے round کریں
    return f"{kanal}-{marla}-{feet}"


@app.post("/export-excel/")
async def export_to_excel(data: RequestData):
    # Prepare static data
    static_data = [
        [""],
        [6, 5, 4, "", 3, 2, 1, ""],
        [
            "اراضی جداگانہ (بعد از تقسیم)",
            "",
            "",
            "",
            "اراضی مشترکہ (قبل از تقسیم)",
            "",
            "",
            "",
        ],
        [
            "نام مالکان جن کے حصہ میں کیفیت بروئے تقسیم دیا گیا",
            "رقبہ",
            "نمبر کھیت مندرجہ شجرہ و گاوں با مندرجہ عکس شجرہ منکحہ سند ہذہ یعنی جدید کھیتؤ ں کا نمبر ",
            "",
            "نام مالکان مشترکہ",
            "رقبہ",
            "نمبر کیفیت مندرجہ شجرہ ",
            "",
        ],
        ["", "", "", "", "", "", "", ""],
    ]

    row_index = 5  # Start filling from row 5

    total_sqft = 0
    wanda_total_sqft = 0
    merge_cells_list = []
    rows_to_style = []

    # Fill taqseemkabad data
    for entry_index, entry in enumerate(data.taqseemkabad):
        wanda_total_sqft = 0
        unique_kilas = set()
        if not entry.details:
           continue  # Skip empty entrie
        # while len(static_data) <= entry_index:
        #     static_data.append([f"ونڈہ نمبر {entry_index + 1}", "", "", "حصص", "", "", "", ""])
        for detail in entry.details:
            wanda_total_sqft += convert_to_sqft(detail.raqbha)
            row_index += 1
            kila_base = ''.join(filter(str.isdigit, str(detail.kila)))
            kila_key = f"{detail.mustatil}_{kila_base}"
            unique_kilas.add(kila_key)

        row = [f"ونڈہ نمبر {entry_index + 1}", wanda_total_sqft, "", ":حصص", "", "", "", ""]
        static_data.append(row)

        rows_to_style.append({
            'index': len(static_data) - 1,  # Keep track of the row index
            'cols': [1, 2]  # Specify the columns that need merging (e.g., column B and C)
        })

        row_index += 1
        wanda_total_sqft = 0 

        grouped_by_malik = defaultdict(list)  
        for detail in entry.details:
            grouped_by_malik[detail.malikName].append(detail)
            wanda_total_sqft += convert_to_sqft(detail.raqbha)


        # Step 2: For each group, add rows
        start_index = len(static_data)  # Where malikName data starts

        for malikName in grouped_by_malik:
            details = grouped_by_malik[malikName]

            for i, detail in enumerate(details):
                name_to_show = malikName if i == 0 else ""  # Only first row gets name

                static_data.append(
                    [
                       name_to_show,
                        detail.raqbha,
                        detail.kila,
                        detail.mustatil,
                        "",
                        "",
                        "",
                        "",
                    ]
                )

        end_index = len(static_data)  # Where malikName data ends

        # Step 1: Get only the relevant slice of names from column A
        names_only = [row[0] for row in static_data[start_index:end_index]]

        # Step 2: Sort column A of only those rows (names first, empty last)
        names_only_sorted = sorted(names_only, key=lambda name: name == "")

        # Step 3: Replace only column A in that same range
        for i, name in enumerate(names_only_sorted):
            static_data[start_index + i][0] = name

            

        row_index += 1
         # میزان ونڈہ کیلکولیشن
        wanda_total_marla = wanda_total_sqft / 272  # Convert total sqft to marlas

        wanda_canal = int(wanda_total_marla // 20)  # 1 Canal = 20 Marlas
        wanda_remaining_marla = wanda_total_marla % 20  # Get remaining marlas after extracting canals

        wanda_marla = int(wanda_remaining_marla)  # Extract whole marlas
        wanda_feet = round((wanda_remaining_marla - wanda_marla) * 272)  # Convert remaining fraction into feet

        wanda_raqbha = f"{wanda_canal}-{wanda_marla}-{wanda_feet}"

        kita_value = str(len(unique_kilas))



        # static_data[entry_index][1] = wanda_total_sqft

        static_data.append([f"{kita_value} قطعہ", wanda_raqbha, "", "میزان ونڈہ", "", "", "", ""])
        # merge_cells_list.append(len(static_data))

        rows_to_style.append({
            'index': len(static_data) - 1,  # Keep track of the row index
            'cols': [1, 2]  # Specify the columns to merge (e.g., column B and C)
        })
        
        print(rows_to_style)

    row_index = 6

    for entry_index, item in enumerate(data.taqseemSaPhala):
        if row_index >= len(static_data):
            static_data.append(["", "", "", "", "", "", "", ""])

        static_data[row_index][4] = item.malikName or ""  # Column E (index 4)
        static_data[row_index][5] = item.totalRaqbha or ""  # Column F (index 5)
        static_data[row_index][6] = item.kilaNumber or ""  # Column G (index 6)
        static_data[row_index][7] = item.mustatil or ""  # Column H (index 7)

        row_index += 1

   # Initialize the total feet
    total_feet = 0

    # Iterate over the rows, starting from row 6 (index 5) up to the last row with data in column F
    for row in static_data[5:]:  # Starting from row 6 (index 5)
        column_f_value = row[5]  # Column F (index 5)

        time_parts = column_f_value.split('-')
        
        if len(time_parts) == 3:
            try:
                canal = float(time_parts[0])  # Canal value
                marla = float(time_parts[1])  # Marla value
                feet = float(time_parts[2])  # Feet value
                
                # Convert Canal to Marla (1 Canal = 272 Marlas)
                canal_in_marla = canal * 20 * 272
                
                # Convert Marla to Feet (1 Marla = 272 Feet)
                marla_in_feet = marla * 272
                
                # Sum up the values in feet (including feet from the Canal, Marla, and direct Feet)
                total_feet += canal_in_marla + marla_in_feet + feet
            except ValueError:
                # Handle the case where the data is not in the expected format
                print(f"Invalid format in row: {row}")
                
    # After the loop, total_feet will contain the sum of all values in column F converted to feet
    print(f"Total sum in feet: {total_feet}")

    

    total_marla = total_feet / 272  # Convert feet to Marlas
    canal = int(total_marla // 20)  # Extract whole canals
    remaining_marla = total_marla % 20  # Get remaining marlas
    marla = int(remaining_marla)  # Extract whole marlas
    remaining_feet = (remaining_marla - marla) * 272  # Convert leftover marlas into feet
    feet = int(remaining_feet)

    totalFeetRounded = int(total_feet)
    print(f"Total sum in feet: {totalFeetRounded}")

 
    totalRaqbha = f"{canal}-{marla}-{feet}"  # Store in variable as a string

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active

    # Add static data to worksheet
    for row in static_data:
        ws.append(row)

    # Apply styling to the first four rows (Headers)
    header_style = Font(bold=True, size=15)  # Bold font for rows 1 and 3
    normal_style = Font(bold=False, size=16)  # Normal font for rows 2 and 4

    for r in range(1, 5):  # Iterate over rows 1 to 4
        for c in range(1, 9):  # Iterate over columns 1 to 8
            cell = ws.cell(row=r, column=c)
        
            if r in [1, 3]:  # Apply bold, center alignment, and wrap text to rows 1 and 3
               cell.font = header_style
               cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
            elif r in [2, 4]:  # Apply center alignment and wrap text (without bold) to rows 2 and 4
                cell.font = normal_style
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Adjust column widths
    column_widths = [32, 10, 10, 10, 33, 10, 7, 7]
    for i, width in enumerate(column_widths):
        ws.column_dimensions[chr(65 + i)].width = width

    # Add logos to the header section (local images)
    logo1_path = os.path.join("assets", "Picture1.png")  # Path to your first logo image
    logo2_path = os.path.join(
        "assets", "Picture2.png"
    )  # Path to your second logo image

    try:
        logo1 = Image(logo1_path)
        logo2 = Image(logo2_path)

        # Resize the logos
        logo1.width = 160
        logo1.height = 0.6 * 160  # Maintain aspect ratio based on width
        logo2.width = 100
        logo2.height = 0.7 * 140  # Maintain aspect ratio based on width

        # Add logo1 to the top-left corner (cell A1)
        logo1.anchor = "A1"
        ws.add_image(logo1)

        # Add logo2 to the top-right corner (cell H1)
        logo2.anchor = "G1"
        ws.add_image(logo2)
    except FileNotFoundError:
        # Handle case where logo files are not found
        pass

    # ws["A1"] = "نقشہ (ج) تقسیم اراضی"  # Replace "A1" with the actual top-left cell
    
    chak = data.chak
    KanonGoyi = data.tehsil
    khataNumber = data.khata
    zila = data.district
    chak_number = data.chak.replace("چک نمبر", "").strip()

    header_text = (
    f"نقشہ (ج) تقسیم اراضی\n\n"
    f"کھاتہ نمبر: {khataNumber} چک/موضع: {chak_number}  "
    f"تحصیل: {KanonGoyi}  ضلع: {zila}"
)

    # Merge cells A1 to H1 and set the text box

    malik_name_column = 5  # Excel-wise it is Column 'E'

    for row in range(6, ws.max_row + 1):
        ws.cell(row=row, column=malik_name_column).alignment = Alignment(wrap_text=True)

        for col in [6, 7, 8]: 
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center',vertical="center", )

    malik_name_column = 1  

    for row in range(6, ws.max_row + 1):
        ws.cell(row=row, column=malik_name_column).alignment = Alignment(wrap_text=True)
        
        for col in [2, 3, 4]:  
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center',vertical="center",)

    ws.merge_cells("A1:H1")
    header_cell = ws["A1"]  # Removed the trailing comma that was making it a tuple
    header_cell.value = header_text
    header_cell.font = Font(size=20, bold=True)
    header_cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    totalFeetRounded = convert_to_sqft(totalRaqbha)



    ws["F6"] = f"{totalRaqbha} میزان کھاتہ "

    ws["F6"].font = Font(size=14, bold=True)
    ws["F6"].alignment = Alignment(horizontal="center", vertical="center")

    ws["E6"] = f"{totalFeetRounded} حصہ"
    ws["E6"].font = Font(size=14, bold=True)
    ws["E6"].alignment = Alignment(horizontal="center", vertical="center")


    start_row = 7
    max_row = ws.max_row

    # ---------- Loop 1: Clean E, F, G, H without preservation ----------
    for col_letter in ["E", "F", "G", "H"]:
        non_empty_values = []

        # Step 1: Collect valid data
        for row in range(start_row, max_row + 1):
            cell_value = ws[f"{col_letter}{row}"].value
            if cell_value not in (None, "", "0-0-0"):
                non_empty_values.append(cell_value)

        # Step 2: Clear entire column
        for row in range(start_row, max_row + 1):
            ws[f"{col_letter}{row}"].value = None

        # Step 3: Write valid data back upward
        for i, value in enumerate(non_empty_values):
            ws[f"{col_letter}{start_row + i}"].value = value

        # Step 4: Fill the remaining rows with empty string
        for j in range(len(non_empty_values), max_row - start_row + 1):
            ws[f"{col_letter}{start_row + j}"].value = ""

        # ---------- Loop 2: Clean A, B, C, D while preserving rows with "میزان ونڈہ" or ":حصص" ----------
        # Step 0: Identify rows that contain "میزان ونڈہ" or ":حصص"
    start_row = 7
    max_row = ws.max_row

    # Step 0: Identify preserved rows
    preserved_rows = set()
    for row in range(start_row, max_row + 1):
        for col in ["B", "C", "D"]:
            val = str(ws[f"{col}{row}"].value).strip() if ws[f"{col}{row}"].value else ""
            if val in ("میزان ونڈہ", ":حصص"):
                preserved_rows.add(row)
                break

    # Step 1: Process columns A–D
    for col_letter in ["B", "C", "D"]:
        non_empty_values = []

        # Collect valid values
        for row in range(start_row, max_row + 1):
            if row in preserved_rows:
               continue
            cell_value = ws[f"{col_letter}{row}"].value
            if cell_value not in (None, "", "0-0-0", "میزان ونڈہ", ":حصص"):
                non_empty_values.append(cell_value)

        # Clear all values (except preserved)
        for row in range(start_row, max_row + 1):
            if row not in preserved_rows:
                ws[f"{col_letter}{row}"].value = None

        # Refill values
        write_row = start_row
        value_index = 0
        while write_row <= max_row:
            if write_row in preserved_rows:
                write_row += 1
                continue
            if value_index < len(non_empty_values):
                ws[f"{col_letter}{write_row}"].value = non_empty_values[value_index]
                value_index += 1
            else:
                ws[f"{col_letter}{write_row}"].value = ""
            write_row += 1

    # Merge columns in row 2
    ws.merge_cells("C2:D2")  # Merging B and C
    ws.merge_cells("G2:H2")  # Merging F and G

    # Merge columns in row 3
    ws.merge_cells("A3:D3")  # Merging A, B, C, D
    ws.merge_cells("E3:H3")  # Merging E, F, G, H

    ws.merge_cells("G4:H5")
    ws.merge_cells("F4:F5")
    ws.merge_cells("E4:E5")


    cell = ws["C4"]
    cell.font = Font(bold=False, size=12)
    ws.merge_cells("C4:D5")
    ws.merge_cells("F6:H6")
    ws.merge_cells("B4:B5")
    ws.merge_cells("A4:A5")

    ws["G6"].font = Font(bold=False, size=14)
    ws["F6"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    # Center align merged cells
    merged_cells = ["B2", "F2", "A3", "E3"]
    for cell in merged_cells:
        ws[cell].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    for row_info in rows_to_style:
        row_index = row_info['index'] + 1  # Convert to 1-based index for openpyxl
       
        ws.merge_cells(f"B{row_index}:C{row_index}")

        cell = ws[f"A{row_index}"]
        cell.font = Font(bold=True, size=14)  # Make text bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Center align text

        cell = ws[f"B{row_index}"]
        cell.font = Font(bold=True, size=14)  # Make text bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Center align text

        cell = ws[f"D{row_index}"]
        cell.font = Font(bold=True, size=14)  # Make text bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Center align text

    ws.row_dimensions[1].height = 160
    ws.row_dimensions[4].height = 90

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    thin = Side(style='thin')
    no_border = Side(style=None) 

    for col in range(6, 9):  # Columns F (6) to H (8)
        cell = f"{get_column_letter(col)}6"
        ws[cell].border = thin_border
    
    # Apply border to each cell in columns A to H
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=8):
        for cell in row:
            cell.border = thin_border

    # Step 2: Loop through each column A to H (1 to 8)
    for col in range(1, 9):
        # Collect all non-empty cells in this column from row 6 onwards
        data_cells = []
        for row in range(6, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                data_cells.append(cell)

            if not data_cells:
                last_cell = ws.cell(row=ws.max_row, column=col)
                last_cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
                print("last_cell")
                continue

            # Define first and last data cell in the column
        first_cell = data_cells[0]
        last_cell = data_cells[-1]


        for cell in data_cells:
            # Borders
            top = thin if cell == first_cell else None
            bottom = thin if cell == last_cell else None
            left = thin
            right = thin

        # Apply border
            cell.border = Border(
                top=top,
               bottom=bottom,
                left=left,
                right=right
            )

    for row_info in rows_to_style:
        row_index = row_info['index'] + 1  # Convert to 1-based index for openpyxl
        ws.merge_cells(f"B{row_index}:C{row_index}")

        # Set font and alignment for A, B, D columns
        # Column A
        cell = ws[f"A{row_index}"]
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border  # Apply border to cell

        # Column B
        cell = ws[f"B{row_index}"]
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border  # Apply border to cell

                # Column B
        cell = ws[f"C{row_index}"]
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border  # Apply border to cell

        # Column D
        cell = ws[f"D{row_index}"]
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border  # Apply border to cell


        cell = ws[f"F6"]
        cell.border = thin_border  # Apply border to cell

        cell = ws[f"E6"]
        cell.border = thin_border  # Apply border to cell





    # Set the page orientation to Portrait
    ws.page_setup.orientation = "portrait"
    ws.page_setup.horizontalCentered = True
    ws.page_setup.paperSize = 5
    ws.print_title_rows = '1:5'
    # ws.row_breaks.append(37) 

    # ws.oddFooter.center.size = 10
    # ws.oddFooter.center.font = "Arial"

    # # Footer text in center (with line below each label)
    # ws.oddFooter.center.text = (
    #     "&B\n\n\n\n\n\n\n\n\n\n\n"  # Adds extra line breaks
    #     "دستخط پٹواری                              دستخط گرداور                                     دستخط ریونیو آفیسر\n"
    #     "__________________________    ________________________    _______________________"
    # )

    empty_rows_to_add = 10  # Number of empty rows to add before the footer
    
    for _ in range(empty_rows_to_add):
        static_data.append([""] * 8)  # Assuming 8 columns in your table

    # Then, add the footer text
    ws.oddFooter.center.size = 10
    ws.oddFooter.center.font = "Arial"
    ws.oddFooter.center.text = (
        "&B\n\n\n\n\n\n\n\n\n\n\n\n"
        "\n "
        
        "\n "
        " __________________________   دستخط پٹواری   _______________________          دستخط گرداور    ________________________          دستخط ریونیو آفیسر"
    )




    # Ensure all columns A to H fit in one page width
    ws.page_setup.fitToWidth = 1  # Force width to fit on one page
    ws.page_setup.fitToHeight = 0  # Allows multiple pages for height if needed

    # Manually set scaling to ensure all columns fit
    ws.page_setup.scale = 80  # Adjust this value (e.g., 60, 70, 90) if necessary

    # Set print area from Column A to H
    ws.print_area = f'A1:H{ws.max_row}'

    # Set narrow margins to give more space for columns
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            existing_font = cell.font
            cell.font = Font(
                name="Jameel Noori Nastaleeq",
                size=existing_font.size,
                bold=existing_font.bold,
                italic=existing_font.italic,
               underline=existing_font.underline,
                color=existing_font.color,
        )

    # Save the file to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Ensure the filename is properly encoded for Unicode characters
    filename = "توزیع_نقشہ.xlsx"
    encoded_filename = urllib.parse.quote(filename)

    # Return the file as a downloadable response with UTF-8 encoded filename
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        },
    )

